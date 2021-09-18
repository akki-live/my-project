import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.colors import BLUE

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}
product_under_ten = {}

print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    if supplier_name in product_per_supplier:
        current_num_product = product_per_supplier[supplier_name]
        product_per_supplier[supplier_name] = current_num_product + 1
    else:
        product_per_supplier[supplier_name] = 1


# calculatioin total value of per suppliers

    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier[supplier_name]
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

# calculation products under 10
    if inventory < 10:
        product_under_ten[product_number] = int(inventory)

# add value to colom 
    inventory_price.value = inventory * price

# add name of 5th colom
inventory_price_colom_name = product_list.cell(1, 5)
inventory_price_colom_name.value = "Total_Value"
inventory_price_colom_name.font = Font(bold=True)


inventory_file.save("akki_inv1.xlsx")

print(product_per_supplier)
print(total_value_per_supplier)
print(product_under_ten)
    

